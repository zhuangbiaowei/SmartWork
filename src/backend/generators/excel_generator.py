import asyncio
from typing import Dict, Any
from pathlib import Path

try:
    from openpyxl import Workbook, Worksheet
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, Reference
except ImportError:
    Workbook = None
    Worksheet = None
    Font = None
    PatternFill = None
    Alignment = None
    Border = None
    Side = None
    BarChart = None
    Reference = None

from .document_generator import DocumentGenerator


class ExcelGenerator(DocumentGenerator):
    """
    Excel document generator using openpyxl.

    Supports formatting, formulas, and charts.
    """

    def __init__(self):
        self.wb = None

    async def generate(
        self, data: Dict[str, Any], output_path: str, **kwargs: Any
    ) -> bool:
        """
        Generate Excel document from data.

        Args:
            data: Dictionary with:
                - title: Spreadsheet title
                - rows: List of rows (each row is a list of values)
                - headers: List of column headers
                - charts: Optional chart data
            output_path: Path to save Excel file

        Returns:
            True if successful, False otherwise
        """
        if Workbook is None:
            print("openpyxl not installed. Install with: pip install openpyxl")
            return False

        try:
            self.wb = Workbook()
            ws = self.wb.active
            ws.title = data.get("title", "Data Report")

            if "headers" in data:
                for col_idx, header in enumerate(data["headers"], 1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.value = header
                    cell.font = Font(bold=True, size=12)
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            if "rows" in data:
                for row_idx, row_data in enumerate(data["rows"], 2):
                    for col_idx, value in enumerate(row_data, 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.value = value

                        if "chart_data" in data:
                            self._add_charts(ws, data["chart_data"])

            self._apply_formats(ws)

            file_path = Path(output_path)
            file_path.parent.mkdir(parents=True, exist_ok=True)

            self.wb.save(output_path)
            return True

        except Exception as e:
            print(f"Failed to generate Excel: {e}")
            return False

    def _add_charts(self, ws: "Worksheet", chart_data: Dict[str, Any]) -> None:
        """
        Add charts to worksheet.

        Args:
            ws: Worksheet to add charts to
            chart_data: Dictionary with chart configuration
        """
        if BarChart is None:
            return None

        if "type" in chart_data:
            chart_type = chart_data["type"]

            if chart_type == "bar":
                chart = BarChart()
                chart.type = chart_data.get("chart_style", "col")

                data_ref = Reference(
                    min_col=chart_data["col_start"],
                    max_col=chart_data["col_end"],
                    min_row=chart_data["row_start"],
                    max_row=chart_data["row_end"],
                )

                chart.add_data(data_ref)

                ws.add_chart(chart_data.get("position", "I12"), chart)

    def _apply_formats(self, ws: "Worksheet") -> None:
        """
        Apply formatting to worksheet.

        Args:
            ws: Worksheet to format
        """
        if Font is None:
            return None

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left", vertical="center")

        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3")

        for cell in ws[1]:
            cell.fill = header_fill
